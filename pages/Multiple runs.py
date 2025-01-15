import streamlit as st
import pandas as pd
import os
import sys
import shutil
from openpyxl import load_workbook
from io import BytesIO
import numpy as np

parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
sys.path.append(parent_dir)
from inp_file_multiple2 import *
from utility import *

st.set_page_config(
    page_title="MKM Input File Generator and Solver",
    page_icon=":coffee:",
)

st.title("MKM Input File Generator and Solver")
st.page_icon = ":coffee:"

def main():
    #os.chdir("D:/projects/mkm_shell/alternative")  # Adjust as per your directory

    # Upload Excel file
    uploaded_file = st.file_uploader("Upload Excel File", type="xlsx")

    # Lists for dropdown selection
    pH_l = [round(x * 0.5, 1) for x in range(0, 29)]  # pH from 0.0 to 14.0
    V_l = [round(x * 0.1, 1) for x in range(-10, 11)]  # V from -1.0 to 1.0

    # Dropdowns for selecting pH and potential
    pH_list = st.multiselect("Select pH Values", pH_l)
    V_list = st.multiselect("Select Potential Values", V_l)

    # Generate Input Files Button
    if st.button("Generate Modified Excel Files"):
        if not uploaded_file:
            st.error("Please upload an Excel file first.")
            return

        for pH in pH_list:
            #parent_folder = os.path.join(os.getcwd(), f"pH_{pH}")
            parent_folder = os.path.join(os.getcwd(), "multiple_run", f"pH_{pH}")

            #print(f"Parent folder for pH={pH}: {parent_folder}")  # Debugging line
            if not os.path.exists(parent_folder):
                os.makedirs(parent_folder)

            for V in V_list:
                children_folder = os.path.join(parent_folder, f"V_{V}")
                #print(f"Children folder for pH={pH}, V={V}: {children_folder}")  # Debugging line

                if not os.path.exists(children_folder):
                    os.makedirs(children_folder)

                try:
                    buffer = modify_excel(children_folder, pH, V, uploaded_file)
                    st.write(f"Excel file modified for pH={pH}, V={V} in {children_folder}")

                    # Display download button with a unique key for Excel file
                    st.download_button(
                        label=f"Download Workbook (pH={pH}, V={V})",
                        data=buffer,
                        file_name=f"inp_file_pH_{pH}_V_{V}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{pH}_{V}"
                    )
                except Exception as e:
                    st.error(f"Error modifying Excel file: {str(e)}")
                    return

    # Run Solver Button
    if st.button("Generate MKM input"):
        if not uploaded_file:
            st.error("Please upload an Excel file first.")
            return

        if not pH_list or not V_list:
            st.error("Please select at least one pH and potential value.")
            return

        for pH in pH_list:
            #parent_folder = os.path.join(os.getcwd(), f"pH_{pH}")
            parent_folder = os.path.join(os.getcwd(), "multiple_run", f"pH_{pH}")

            for V in V_list:
                children_folder = os.path.join(parent_folder, f"V_{V}")
                input_file_path = os.path.join(children_folder, "inp_file.xlsx")

                if not os.path.exists(input_file_path):
                    st.write(input_file_path)
                    st.error(f"Input file not found for pH={pH}, V={V}. Generate files first.")
                    return

                try:
                    inp_file_gen_multiple(input_file_path, children_folder)  # Call the function to generate .mkm files
                    mkm_file_path = os.path.join(children_folder, "input_file.mkm")
                    if os.path.exists(mkm_file_path):
                        st.success(f"Solver successfully generated files for pH={pH}, V={V}. .mkm file found: {mkm_file_path}")

                        # Display download button for .mkm file
                        with open(mkm_file_path, "rb") as mkm_file:
                            mkm_file_data = mkm_file.read()

                        st.download_button(
                            label=f"Download MKM Input File (pH={pH}, V={V})",
                            data=mkm_file_data,
                            file_name=f"input_file_pH_{pH}_V_{V}.mkm",
                            mime="text/plain",
                            key=f"download_mkm_{pH}_{V}"
                        )

                    else:
                        st.error(f".mkm file not found for pH={pH}, V={V}.")
                except Exception as e:
                    st.error(f"Error running solver for pH={pH}, V={V}: {str(e)}")
                    st.write(f"Detailed Exception: {repr(e)}")  # This will print the full exception message
                    raise e

                    return

    if st.button("Run Solver for All Files"):
        if not pH_list or not V_list:
            st.error("Please select at least one pH and potential value.")
            return

        all_success = True  # To track overall success
        for pH in pH_list:
            #parent_folder = os.path.join(os.getcwd(), f"pH_{pH}")
            parent_folder = os.path.join(os.getcwd(), "multiple_run", f"pH_{pH}")

            for V in V_list:
                children_folder = os.path.join(parent_folder, f"V_{V}")
                input_file_path = os.path.join(children_folder, "input_file.mkm")

                if not os.path.exists(input_file_path):
                    st.error(f".mkm file not found for pH={pH}, V={V}. Generate files first.")
                    all_success = False
                    continue

                try:
                    # Run the executable for the input file
                    result_message, success = run_executable(input_file_path)
                    if success:
                        st.success(f"Solver successfully ran for pH={pH}, V={V}: {result_message}")
                        #coverage()  # If there's additional functionality, include it here
                        #plot_coverage_data([pH], [V])  # Plot for this particular pH and V combination
                        coverage()

                    else:
                        st.error(f"Solver failed for pH={pH}, V={V}: {result_message}")
                        all_success = False
                except Exception as e:
                    st.error(f"Error running solver for pH={pH}, V={V}: {str(e)}")
                    all_success = False

        if all_success:
            st.success("Solver ran successfully for all files.")
        else:
            st.warning("Solver encountered errors for some files.")


def modify_excel(children_folder, pH, potential, uploaded_file):
    try:
        # Load the uploaded Excel workbook
        workbook = load_workbook(filename=uploaded_file)

        if "Local Environment" not in workbook.sheetnames:
            st.error("Sheet 'Local Environment' not found.")
            return

        # Access the "Local Environment" sheet
        sheet = workbook["Local Environment"]
        sheet["B2"].value = potential
        sheet["C2"].value = pH

        # Save updated workbook to BytesIO buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Save to a file and copy to children folder
        with open("inp_file.xlsx", "wb") as f:
            f.write(buffer.read())
        shutil.copy("inp_file.xlsx", children_folder)

        # Return buffer for download
        buffer.seek(0)
        return buffer
    except Exception as e:
        st.error(f"Error: {str(e)}")
        raise e


if __name__ == "__main__":
    main()
