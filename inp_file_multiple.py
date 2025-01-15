import os
import streamlit as st
import pandas as pd
import numpy as np
from mkm_parameters import *
import xlwings as xw
import pythoncom
from openpyxl import load_workbook

def read_formulas(file_name, sheet_name, column_name):
    """
    Recalculates the formulas in an Excel file, saves the file, and returns the dataframe with the updated values.
    
    Args:
    file_name (str): Path to the Excel file.
    sheet_name (str): Name of the sheet to process.
    column_name (str): The column name to focus on in the resulting DataFrame.
    
    Returns:
    pd.DataFrame: DataFrame with the computed values from the specified column.
    """
    try:
        # Initialize COM thread for Excel (important for multi-threaded environments like Streamlit)
        pythoncom.CoInitialize()
        # Open Excel file
        app = xw.App(visible=False)  # Open Excel in the background
        wb = xw.Book(file_name)
        sheet = wb.sheets[sheet_name]

        # Recalculate all formulas in the workbook
        wb.app.calculation = 'automatic'
        wb.save(file_name)  # Save the file after recalculation
        wb.close()
        app.quit()

        # Read the recalculated file with pandas
        df = pd.read_excel(file_name, engine="openpyxl", sheet_name=sheet_name)
        
        # Optionally, focus on the specified column if needed
        if column_name in df.columns:
            return df[[column_name]]  # Return only the specified column
        else:
            raise ValueError(f"Column '{column_name}' not found in the sheet.")
    
    except Exception as e:
        st.write(f"Error: {e}")
        print(f"Error: {e}")
        return None 
    
import xlwings as xw
import pandas as pd
from openpyxl import load_workbook

import xlwings as xw
import pythoncom
from openpyxl import load_workbook

def force_recalculate(file_name):
    try:
        pythoncom.CoInitialize()  # Initialize COM
        
        app = xw.App(visible=False)
        wb = app.books.open(file_name)
        wb.app.calculate()  # Force calculation of all formulas
        wb.save()           # Save the file after calculation
        wb.close()
        app.quit()
    except Exception as e:
        print(f"Error in recalculation: {e}")
    finally:
        pythoncom.CoUninitialize()  # Uninitialize COM

def read_formula(file_name, sheet_name, column_name):
    force_recalculate(file_name)  # Ensure all dependent formulas are recalculated

    # Now load the workbook with data_only=True to get the calculated values
    workbook = load_workbook(file_name, data_only=True)

    # Access the specified sheet
    sheet = workbook[sheet_name]

    # Get the maximum row and column number
    max_row = sheet.max_row
    max_column = sheet.max_column

    # Read the data dynamically based on the max_row and max_column
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        data.append([cell.value for cell in row])

    # Use the first row as column headers and the remaining rows as data
    df = pd.DataFrame(data[1:], columns=data[0])

    # Filter the DataFrame to include only the specified column
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' not found in the sheet.")
    
    # Return the DataFrame with the requested column
    return df[column_name]


def inp_file_gen_multiple(uploaded_file,children_folder):
        if uploaded_file:
            try:
                data = pd.read_excel(uploaded_file)
            except Exception as e:
                st.error(f"Error Loading Data: {str(e)}")
                return
            try:
                data1 = pd.read_excel(uploaded_file, sheet_name="Reactions")
                df1 = data1
            except Exception as e:
                st.error(f"Error reading Reactions sheet: {str(e)}")
                return
            
            try:
                data2 = pd.read_excel(uploaded_file, sheet_name="Local Environment")
                df2 = data2
            except Exception as e:
                st.error(f"Error reading Local Environment sheet: {str(e)}")
                return
            
            try:
                data3 = pd.read_excel(uploaded_file, sheet_name="Input-Output Species")
                df3 = data3
            except Exception as e:
                st.error(f"Error reading Input-output sheet: {str(e)}")
                return
        # Extract necessary data from the dataframes
            try:
                global pH_list, V_list, gases, rxn, concentrations, Ea, Eb, P
                pH_list = df2["pH"].tolist()
                V_list = df2["V"][0]
                gases = df3["Species"].tolist()
                concentrations = read_formula(uploaded_file,'Input-Output Species', 'Input MKMCXX')
                rxn = df1["Reactions"]
                Ea = read_formula(uploaded_file,'Reactions', 'G_f')
                Eb = read_formula(uploaded_file,'Reactions', 'G_b')
                P = df2["Pressure"][0]
                st.write("Parameters extracted successfully!")
            except Exception as e:
                st.error(f"Error extracting parameters: {str(e)}")
                return     

            try:
                global adsorbates, activity, Reactant1, Reactant2, Reactant3, Product1, Product2, Product3
                Reactant1 = []
                Reactant2 = []
                Reactant3 = []
                Product1 = []
                Product2 = []
                Product3 = []
                adsorbates = []

                for i in range(len(rxn)):
                    Reactant1.append("{" + rxn[i].split("→")[0].split("+")[0].strip() + "}")
                    if len(rxn[i].split("→")[0].split("+")) == 3:
                        Reactant2.append("{" + rxn[i].split("→")[0].split("+")[1].strip() + "}")
                        Reactant3.append("{" + rxn[i].split("→")[0].split("+")[2].strip() + "}")
                    elif len(rxn[i].split("→")[0].split("+")) == 2:
                        Reactant2.append("{" + rxn[i].split("→")[0].split("+")[1].strip() + "}")
                        Reactant3.append("")
                    else:
                        Reactant2.append("")
                        Reactant3.append("")

                    Product1.append("{" + rxn[i].split("→")[1].split("+")[0].strip() + "}")
                    if len(rxn[i].split("→")[1].split("+")) == 3:
                        Product2.append("{" + rxn[i].split("→")[1].split("+")[1].strip() + "}")
                        Product3.append("{" + rxn[i].split("→")[1].split("+")[2].strip() + "}")
                    elif len(rxn[i].split("→")[1].split("+")) == 2:
                        Product2.append("{" + rxn[i].split("→")[1].split("+")[1].strip() + "}")
                        Product3.append("")
                    else:
                        Product2.append("")
                        Product3.append("")

                for index in Reactant1:
                    if "*" in index and index.strip("{").strip("}") not in adsorbates:
                        adsorbates.append(index.strip("{").strip("}"))

                for index in Reactant2:
                    if "*" in index and index.strip("{").strip("}") not in adsorbates:
                        adsorbates.append(index.strip("{").strip("}"))

                for index in Product1:
                    if "*" in index and index.strip("{").strip("}") not in adsorbates:
                        adsorbates.append(index.strip("{").strip("}"))

                for index in Product2:
                    if "*" in index and index.strip("{").strip("}") not in adsorbates:
                        adsorbates.append(index.strip("{").strip("}"))

                # Check if '*' exists in the list before removing it
                if "*" in adsorbates:
                    adsorbates.remove("*")

                activity = np.zeros(len(adsorbates))

            except Exception as e:
                st.error(f"Error generating input file: {str(e)}")
    

        inp_file_path=os.path.join(children_folder, 'input_file.mkm')
        inp_file = open(inp_file_path, 'w')
        inp_file.write('&compounds\n\n')
        inp_file.write("#gas-phase compounds\n\n#Name; isSite; concentration\n\n")
        for compound,concentration in zip(gases,concentrations):
            inp_file.write("{:<15}; 0; {}\n".format(compound,concentration))

        inp_file.write("\n\n#adsorbates\n\n#Name; isSite; activity\n\n")   
        for compound,concentration in zip(adsorbates,activity):
            inp_file.write("{:<15}; 1; {}\n".format(compound,concentration))

        inp_file.write("\n#free sites on the surface \n\n")
        inp_file.write("#Name; isSite; activity\n\n")   
        inp_file.write("*; 1; {}\n\n".format(1.0))    

        inp_file.write('&reactions\n\n')
        pre_exp=6.21e12
        
        for j in range(len(rxn)):
            if Reactant3[j]!="":
                line = "AR; {:<15} + {:<15} + {:<5} => {:<15}{:<15};{:<10.2e} ;  {:<10.2e} ;  {:<10} ;  {:<10} \n".format(Reactant1[j],Reactant2[j],Reactant3[j],Product1[j],Product2[j],pre_exp, pre_exp, Ea[j],Eb[j] )   
            elif Product3[j]!="":
                line = "AR; {:<15} + {:<14}  => {:<10} + {:<15} + {:<7};{:<10.2e} ;  {:<10.2e} ;  {:<10} ;  {:<10} \n".format(Reactant1[j],Reactant2[j],Product1[j],Product2[j],Product3[j],pre_exp, pre_exp, Ea[j],Eb[j] )     
            elif  Reactant2[j]!="" and Product2[j]!="":
                line = "AR; {:<15} + {:<15} => {:<15} + {:<20};{:<10.2e} ;  {:<10.2e} ;  {:<10} ;  {:<10} \n".format(Reactant1[j],Reactant2[j],Product1[j],Product2[j],pre_exp, pre_exp, Ea[j],Eb[j] )
            elif  Reactant2[j]=="" and Product2[j]!="":
                line = "AR; {:<15} {:<17} => {:<15} + {:<20};{:<10.2e} ;  {:<10.2e} ;  {:<10} ;  {:<10} \n".format(Reactant1[j],"",Product1[j],Product2[j],pre_exp, pre_exp, Ea[j],Eb[j] )
            elif Reactant2[j]!="" and Product2[j]=="":
                line = "AR; {:<15} + {:<15} => {:<15}{:<23};{:<10.2e} ;  {:<10.2e} ;  {:<10} ;  {:<10} \n".format(Reactant1[j],Reactant2[j],Product1[j],"",pre_exp, pre_exp, Ea[j],Eb[j] )
            elif Reactant2[j]=="" and Product2[j]=="":
                line = "AR; {:<15} {:<17} => {:<15}{:<23};{:<10.2e} ;  {:<10.2e} ;  {:<10} ;  {:<10} \n".format(Reactant1[j],"",Product1[j],"",pre_exp, pre_exp, Ea[j],Eb[j] )
            
            inp_file.write(line)    
        inp_file.write("\n\n&settings\nTYPE = SEQUENCERUN\nPRESSURE = {}".format(P))
        inp_file.write("\nPOTAXIS=1\nDEBUG=0\nNETWORK_RATES=1\nNETWORK_FLUX=1\nUSETIMESTAMP=0")
        inp_file.write('\n\n&runs\n')
        inp_file.write("# Temp; Potential;Time;AbsTol;RelTol\n")
        line2 = "{:<5};{:<5};{:<5.2e};{:<5};{:<5}".format(Temp,V_list,Time,Abstol,Reltol)
        inp_file.write(line2)   
        inp_file.close()  
        st.write(f"Input file successfully created at {os.getcwd()}")

