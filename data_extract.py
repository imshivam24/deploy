def data_extract(pH,V,inp_path):
    import pandas as pd
    import os
    import numpy as np
    import xlwings as xw

    ## Modify the excel based on the input file
    from openpyxl import load_workbook

    # Load the Excel workbook
    print(os.getcwd())
    workbook = load_workbook(filename=r"../../input.xlsx")

    # Select the sheet named 'local environment'
    sheet = workbook['Local Environment']

    # Find the index of the 'pH' column
    ph_column_index = None
    for cell in sheet[1]:  # Assuming the first row contains headers
        if cell.value == 'pH':
            ph_column_index = cell.column  # Get the column index of 'pH'
            break

    if ph_column_index is not None:
        # Iterate through rows in the 'pH' column (starting from row 2)
        for row in range(2, sheet.max_row + 1):  # Starting from row 2 to skip header
            current_value = sheet.cell(row=row, column=ph_column_index).value  # Accessing 'pH' column
            if current_value is not None:
                # Modify value 
                new_value = pH  # Assign new value
                sheet.cell(row=row, column=ph_column_index).value = new_value  # Assign new value back to cell

        # Save the modified workbook
        workbook.save(filename='input_data.xlsx')
    else:
        print("The 'pH' column was not found.")

    # Find the index of the 'V' column
    V_column_index = None
    for cell in sheet[1]:  # Assuming the first row contains headers
        if cell.value == 'V':
            V_column_index = cell.column  # Get the column index of 'V'
            break

    if V_column_index is not None:
        # Iterate through rows in the 'V' column (starting from row 2)
        for row in range(2, sheet.max_row + 1):  # Starting from row 2 to skip header
            current_value = sheet.cell(row=row, column=V_column_index).value  # Accessing V column
            if current_value is not None:
                # Modify value 
                new_value = V # Assign new value
                sheet.cell(row=row, column=V_column_index).value = new_value  # Assign new value back to cell

        # Save the modified workbook
        workbook.save(filename='input_data.xlsx')
    else:
        print("The 'pH' column was not found.") 

    import xlwings as xw

    def read_formulas(inp_path, sheet_name, column_name):
        """Read all values under the specified column name in the specified worksheet."""
        with xw.Book(inp_path) as wb:  # Automatically closes when exiting this block
            # Access the specified worksheet
            sheet = wb.sheets[sheet_name]
            
            # Find the header row
            header_row = sheet.range('1:1').value  # Assuming headers are in the first row
            
            # Get the index of the specified column
            if column_name in header_row:
                column_index = header_row.index(column_name) + 1  # +1 because index is zero-based
                
                # Read all values in the specified column (starting from row 2 to skip header)
                column_values = sheet.range((2, column_index), (sheet.cells.last_cell.row, column_index)).value
                
                # Convert to list and filter out None values
                values_list = [value for value in column_values if value is not None]
                return values_list
            else:
                print(f"Header '{column_name}' not found.")
                return [] 
      
                
    Ea = read_formulas(inp_path,'Reactions', 'G_f')
    Eb = read_formulas(inp_path,'Reactions', 'G_b')
    concentrations = read_formulas(inp_path,'Input-Output Species', 'Input MKMCXX')
    rxn=pd.read_excel(inp_path,sheet_name='Reactions')["Reactions"]
    #Ea=pd.read_excel(inp_path,sheet_name='Reactions')["G_f"]
    #Eb=pd.read_excel(inp_path,sheet_name='Reactions')["G_b"]
    V=pd.read_excel(inp_path,sheet_name='Local Environment')["V"][0]
    pH=pd.read_excel(inp_path,sheet_name='Local Environment')["pH"][0]
    P=pd.read_excel(inp_path,sheet_name='Local Environment')["Pressure"][0]
    gases=pd.read_excel(inp_path,sheet_name='Input-Output Species')["Species"].to_list()
    #concentrations=pd.read_excel(inp_path,sheet_name='Input-Output Species')["Concentration"].to_list()


    Reactant1=[]
    Reactant2=[]
    Reactant3=[]
    Product1=[]
    Product2=[]
    Product3=[]
    adsorbates=[]
    for i in range(len(rxn)):
        Reactant1.append("{"+rxn[i].strip().split("→")[0].split("+")[0].strip()+"}")
        if(len(rxn[i].strip().split("→")[0].split("+"))==3):
            Reactant2.append("{"+rxn[i].strip().split("→")[0].split("+")[1].strip()+"}")
            Reactant3.append("{"+rxn[i].strip().split("→")[0].split("+")[2].strip()+"}")

        elif (len(rxn[i].strip().split("→")[0].split("+"))==2):
            Reactant2.append("{"+rxn[i].strip().split("→")[0].split("+")[1].strip()+"}")
            Reactant3.append("")
        else:
            Reactant2.append("")
            Reactant3.append("")    

        Product1.append("{"+rxn[i].strip().split("→")[1].split("+")[0].strip()+"}")
        if(len(rxn[i].strip().split("→")[1].split("+"))==3):
            Product2.append("{"+rxn[i].strip().split("→")[1].split("+")[1].strip()+"}")
            Product3.append("{"+rxn[i].strip().split("→")[1].split("+")[2].strip()+"}")
        elif(len(rxn[i].strip().split("→")[1].split("+"))==2):
            Product2.append("{"+rxn[i].strip().split("→")[1].split("+")[1].strip()+"}")
            Product3.append("")  
        else:
            Product2.append("")
            Product3.append("")

    for index in Reactant1:
        if "*" in index and index.strip("{").strip("}") not in adsorbates:
            adsorbates.append(index.strip("{").strip("}"))

    for index in Reactant2:
        if "*" in index and index.strip("{").strip("}") not in adsorbates:
            adsorbates.append(index.strip("{").strip("}"))

    for index in Product1:
        if "*" in index and index.strip("{").strip("}") not in adsorbates:
            adsorbates.append(index.strip("{").strip("}"))

    for index in Product2:
        if "*" in index and index.strip("{").strip("}") not in adsorbates:
            adsorbates.append(index.strip("{").strip("}"))       

    adsorbates.remove("*")
    activity=np.zeros(len(adsorbates))
    return gases,concentrations,adsorbates,activity,Reactant1,Reactant2,Reactant3,Product1,Product2,Product3,Ea,Eb,P,rxn


